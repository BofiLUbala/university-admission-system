from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions, status
from django.db.models import Count, Sum
from admissions.models import AdmissionApplication
from payments.models import Payment
from faculties.models import Faculty
from accounts.permissions import IsAdminUser
from django.http import HttpResponse
import openpyxl

class AdminStatsView(APIView):
    permission_classes = (permissions.IsAuthenticated, IsAdminUser)

    def get(self, request):
        applications = AdmissionApplication.objects.all()
        total_apps = applications.count()
        approved = applications.filter(status='APPROVED').count()
        rejected = applications.filter(status='REJECTED').count()
        pending = applications.filter(status__in=['SUBMITTED', 'UNDER_REVIEW']).count()
        drafts = applications.filter(status='DRAFT').count()
        
        # Calculate rates
        admission_rate = (approved / total_apps * 100) if total_apps > 0 else 0
        
        # Simulated/Actual Revenue
        successful_payments = Payment.objects.filter(status='SUCCESSFUL')
        total_revenue = successful_payments.aggregate(total=Sum('amount'))['total'] or 0.0
        
        # Faculty distribution
        faculty_dist = AdmissionApplication.objects.values('faculty__name').annotate(count=Count('id')).order_by('-count')
        faculty_stats = []
        for item in faculty_dist:
            if item['faculty__name']:
                faculty_stats.append({
                    'name': item['faculty__name'],
                    'value': item['count']
                })

        # Province distribution
        province_dist = AdmissionApplication.objects.values('province').annotate(count=Count('id')).order_by('-count')
        province_stats = []
        for item in province_dist:
            if item['province']:
                province_stats.append({
                    'name': item['province'],
                    'value': item['count']
                })

        return Response({
            'summary': {
                'total_applications': total_apps,
                'approved': approved,
                'rejected': rejected,
                'pending': pending,
                'drafts': drafts,
                'admission_rate': round(admission_rate, 2),
                'total_revenue': float(total_revenue)
            },
            'faculty_distribution': faculty_stats,
            'province_distribution': province_stats
        })

class ExportApplicationsExcelView(APIView):
    permission_classes = (permissions.IsAuthenticated, IsAdminUser)

    def get(self, request):
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Candidatures UPK"

        # Headers
        headers = [
            "ID Candidature", "Nom Complet", "Genre", "Date de Naissance", 
            "Nationalité", "Province", "Adresse", "Téléphone", "Email", 
            "École d'Origine", "Pourcentage", "Faculté", "Département", 
            "Statut", "Année Académique", "Date de Création"
        ]
        ws.append(headers)

        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="0D3B66", end_color="0D3B66", fill_type="solid")

        # Fetch applications
        applications = AdmissionApplication.objects.all().select_related('faculty', 'department').order_by('-created_at')

        # Rows
        for app in applications:
            ws.append([
                app.id,
                app.full_name,
                app.get_gender_display(),
                app.date_of_birth.strftime("%Y-%m-%d") if app.date_of_birth else "",
                app.nationality,
                app.province,
                app.address,
                app.phone_number,
                app.email,
                app.previous_school,
                float(app.percentage_obtained),
                app.faculty.name if app.faculty else "N/A",
                app.department.name if app.department else "N/A",
                app.get_status_display(),
                app.academic_year,
                app.created_at.strftime("%Y-%m-%d %H:%M") if app.created_at else ""
            ])

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        # Build Response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=candidatures_ulk.xlsx"
        
        wb.save(response)
        return response
